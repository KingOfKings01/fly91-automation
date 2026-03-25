from fpdf import FPDF
import os
import tempfile
import openpyxl
from num2words import num2words
import warnings
import datetime

# Suppress warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# Config
EXCEL_FILE = 'Baggage invoicing.xlsm'
OUTPUT_BASE_DIR = 'Invoices_Output_Final'
MEDIA_DIR = os.path.join(os.path.dirname(__file__), 'media')
UPLOAD_MEDIA_DIR = os.path.join(tempfile.gettempdir(), 'fly91_media')

LOGO_PATH = os.path.join(MEDIA_DIR, 'image1.png')
# Seal and Sign: Check UPLOAD_MEDIA_DIR (/tmp) first, then repo defaults
SEAL_PATH = os.path.join(UPLOAD_MEDIA_DIR, 'image2.png')
if not os.path.exists(SEAL_PATH): SEAL_PATH = os.path.join(MEDIA_DIR, 'image2.png')

SIGN_PATH = os.path.join(UPLOAD_MEDIA_DIR, 'image3.png')
if not os.path.exists(SIGN_PATH): SIGN_PATH = os.path.join(MEDIA_DIR, 'image3.png')

FOOTER_IMAGE_PATH = os.path.join(MEDIA_DIR, 'image4.jpg')

def excel_to_dict_list(path, sheet_name):
    """Lighter replacement for pd.read_excel().to_dict('records')"""
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        if sheet_name not in wb.sheetnames:
            return []
        sheet = wb[sheet_name]
        
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):
                data.append(dict(zip(headers, row)))
        return data
    except Exception as e:
        print(f"Excel read error: {e}")
        return []

def is_empty(val):
    if val is None: return True
    if str(val).lower() == 'nan': return True
    if str(val).strip() == '': return True
    return False

class ProfessionalInvoice(FPDF):
    def __init__(self, *args, **kwargs):
        super().__init__(orientation='L', unit='mm', format='A4', **kwargs)
        self.red_color = (237, 28, 36)
        self.grey_line = (200, 200, 200)
        self.supplier_address_str = ""
        self.supplier_gstin_str = ""

    def header(self):
        if os.path.exists(LOGO_PATH):
            self.image(LOGO_PATH, 10, 8, 45)
        self.set_y(15)
        self.set_font('helvetica', 'B', 11)
        self.set_text_color(0, 0, 0)
        self.cell(0, 5, 'TAX INVOICE', 0, 1, 'C')
        self.set_font('helvetica', '', 7)
        self.cell(0, 4, '(Original for Recipient)', 0, 1, 'C')
        self.ln(5)

    def footer(self):
        if os.path.exists(FOOTER_IMAGE_PATH):
            self.image(FOOTER_IMAGE_PATH, 28.5, 175, 240)
        self.set_y(-10)
        self.set_font('helvetica', '', 7)
        self.set_text_color(150, 150, 150)

def clean_filename(s):
    if not isinstance(s, str): return str(s)
    for c in r'\/:*?"<>|':
        s = s.replace(c, '')
    return s.strip()

def format_num(val):
    try:
        f = float(val) if not is_empty(val) else 0.0
        if f == 0: return "-"
        return "{:,.2f}".format(float(round(f)))
    except:
        return "-"

def number_to_words_indian(num):
    try:
        val = float(num) if not is_empty(num) else 0.0
        if val == 0: return "Zero Only"
        return num2words(val, lang='en_IN').title().replace("-", " ").replace(",", "") + " Only"
    except:
        return ""

def generate_kind_pdf(data, output_path, seal_pos=None, sign_pos=None):
    pdf = ProfessionalInvoice()
    pdf.supplier_address_str = data.get('supplier_address', '')
    pdf.supplier_gstin_str = data.get('supplier_gstin', '')
    pdf.add_page()
    
    line_h = 4.5
    title_fs = 8
    val_fs = 7.5
    
    pdf.set_y(32)
    pdf.set_font('helvetica', 'B', title_fs)
    pdf.cell(30, line_h, "Invoice No")
    pdf.set_font('helvetica', '', title_fs)
    pdf.cell(5, line_h, ":")
    pdf.set_font('helvetica', '', val_fs)
    pdf.cell(100, line_h, str(data['invoice_no']), 0, 0)
    
    pdf.set_x(150)
    pdf.set_font('helvetica', 'B', 9)
    pdf.cell(0, line_h, 'JUST UDO AVIATION PVT LTD', 0, 1, 'R')
    
    pdf.set_font('helvetica', 'B', title_fs)
    pdf.cell(30, line_h, "Invoice Date")
    pdf.set_font('helvetica', '', title_fs)
    pdf.cell(5, line_h, ":")
    pdf.set_font('helvetica', '', val_fs)
    pdf.cell(100, line_h, str(data['invoice_date']), 0, 0)
    
    pdf.set_font('helvetica', '', 6.5)
    pdf.set_x(217)
    pdf.multi_cell(70, 2.5, pdf.supplier_address_str, 0, 'R')
    
    pdf.ln(4)
    pdf.set_font('helvetica', 'B', title_fs)
    pdf.cell(30, line_h, "Passenger Name")
    pdf.set_font('helvetica', '', title_fs)
    pdf.cell(5, line_h, ":")
    pdf.set_font('helvetica', '', val_fs)
    pdf.cell(100, line_h, str(data['passenger_name']), 0, 0)
    
    route_x = 155 
    pdf.set_x(route_x) 
    pdf.set_font('helvetica', 'B', title_fs)
    pdf.cell(35, line_h, "From")
    pdf.set_font('helvetica', '', title_fs)
    pdf.cell(5, line_h, ":")
    pdf.set_font('helvetica', '', val_fs)
    pdf.cell(0, line_h, str(data['origin']), 0, 1)

    pdf.set_font('helvetica', 'B', title_fs)
    pdf.cell(30, line_h, "PNR No")
    pdf.set_font('helvetica', '', title_fs)
    pdf.cell(5, line_h, ":")
    pdf.set_font('helvetica', '', val_fs)
    pdf.cell(50, line_h, str(data['pnr_no']), 0, 0)
    
    pdf.set_x(100) 
    pdf.set_font('helvetica', 'B', title_fs)
    pdf.cell(20, line_h, "Flight No")
    pdf.set_font('helvetica', '', title_fs)
    pdf.cell(5, line_h, ":")
    pdf.set_font('helvetica', '', val_fs)
    pdf.cell(30, line_h, str(data['flight_no']), 0, 0)
    
    pdf.set_x(route_x)
    pdf.set_font('helvetica', 'B', title_fs)
    pdf.cell(35, line_h, "To")
    pdf.set_font('helvetica', '', title_fs)
    pdf.cell(5, line_h, ":")
    pdf.set_font('helvetica', '', val_fs)
    pdf.cell(0, line_h, str(data['destination']), 0, 1)

    pdf.set_x(route_x)
    pdf.set_font('helvetica', 'B', title_fs)
    pdf.cell(35, line_h, "Place of Supply")
    pdf.set_font('helvetica', '', title_fs)
    pdf.cell(5, line_h, ":")
    pdf.set_font('helvetica', '', val_fs)
    pdf.cell(40, line_h, str(data['place_of_supply']), 0, 0)
    
    pdf.set_x(225)
    pdf.set_font('helvetica', 'B', title_fs)
    pdf.cell(30, line_h, "GSTIN of Supplier")
    pdf.set_font('helvetica', '', title_fs)
    pdf.cell(5, line_h, ":")
    pdf.set_font('helvetica', '', val_fs)
    pdf.cell(0, line_h, f"{pdf.supplier_gstin_str}", 0, 1)
    
    pdf.ln(3)
    pdf.set_font('helvetica', 'B', title_fs)
    pdf.cell(30, line_h, "GSTIN of Customer")
    pdf.set_font('helvetica', '', title_fs)
    pdf.cell(5, line_h, ":")
    pdf.set_font('helvetica', '', val_fs)
    pdf.cell(0, line_h, str(data['customer_gstin']), 0, 1)

    pdf.set_font('helvetica', 'B', title_fs)
    pdf.cell(30, line_h, "Customer Name")
    pdf.set_font('helvetica', '', title_fs)
    pdf.cell(5, line_h, ":")
    pdf.set_font('helvetica', '', val_fs)
    pdf.cell(0, line_h, str(data['customer_name']), 0, 1)

    pdf.set_font('helvetica', 'B', title_fs)
    pdf.cell(30, line_h, "Customer Address")
    pdf.set_font('helvetica', '', title_fs)
    pdf.cell(5, line_h, ":")
    pdf.set_font('helvetica', '', 7)
    pdf.multi_cell(0, 3.5, str(data['customer_address']))
    
    pdf.set_y(pdf.get_y() + 2)
    pdf.set_font('helvetica', 'B', 8)
    pdf.cell(275, 5, "Currency : INR", 0, 1, 'R')
    
    w = [60, 18, 22, 18, 18, 12, 15, 12, 15, 12, 15, 12, 15, 31]
    y_hdr = pdf.get_y()
    pdf.set_font('helvetica', 'B', 8)
    pdf.cell(w[0], 12, "Description", 1, 0, 'C')
    pdf.cell(w[1], 12, "SAC Code", 1, 0, 'C')
    pdf.cell(w[2], 12, "Taxable Value", 1, 0, 'C')
    pdf.cell(w[3], 12, "", 1, 0, 'C') 
    pdf.cell(w[4], 12, "Total", 1, 0, 'C')
    pdf.cell(w[5]+w[6], 6, "IGST", 1, 0, 'C')
    pdf.cell(w[7]+w[8], 6, "CGST", 1, 0, 'C')
    pdf.cell(w[9]+w[10], 6, "SGST/UTGST", 1, 0, 'C')
    pdf.cell(w[11]+w[12], 6, "CESS", 1, 0, 'C')
    pdf.cell(w[13], 6, "Total (Incl Taxes)", 1, 1, 'C')
    
    pdf.set_xy(10 + w[0] + w[1] + w[2] + w[3] + w[4], y_hdr + 6)
    pdf.set_font('helvetica', 'B', 6)
    for _ in range(4):
        pdf.cell(w[5], 6, "TAX%", 1, 0, 'C')
        pdf.cell(w[6], 6, "Amount", 1, 0, 'C')
    pdf.set_font('helvetica', 'B', 7)
    pdf.cell(w[13], 6, "Amount", 1, 1, 'C')
    
    pdf.set_font('helvetica', 'B', 6.5)
    pdf.set_xy(10+w[0]+w[1]+w[2], y_hdr + 1)
    pdf.multi_cell(w[3], 3.2, "Non Taxable/\nExempted\nvalue", 0, 'C')
    pdf.set_y(y_hdr + 12)
    
    pdf.set_font('helvetica', '', 9)
    pdf.cell(w[0], 10, str(data['description']), 1, 0, 'L')
    pdf.cell(w[1], 10, str(data['sac_code']), 1, 0, 'C')
    pdf.cell(w[2], 10, format_num(data['taxable_value']), 1, 0, 'R')
    pdf.cell(w[3], 10, format_num(data.get('non_taxable', 0)), 1, 0, 'R')
    pdf.cell(w[4], 10, format_num(data['taxable_value']), 1, 0, 'R')
    
    igst = float(data['igst']) if not is_empty(data['igst']) else 0.0
    if igst > 0:
        pdf.cell(w[5], 10, "5%", 1, 0, 'C')
        pdf.cell(w[6], 10, format_num(data['igst']), 1, 0, 'R')
        pdf.cell(w[7], 10, "-", 1, 0, 'C')
        pdf.cell(w[8], 10, "-", 1, 0, 'R')
        pdf.cell(w[9], 10, "-", 1, 0, 'C')
        pdf.cell(w[10], 10, "-", 1, 0, 'R')
    else:
        pdf.cell(w[5], 10, "0%", 1, 0, 'C')
        pdf.cell(w[6], 10, "-", 1, 0, 'R')
        pdf.cell(w[7], 10, "2.5%", 1, 0, 'C')
        pdf.cell(w[8], 10, format_num(data['cgst']), 1, 0, 'R')
        pdf.cell(w[9], 10, "2.5%", 1, 0, 'C')
        pdf.cell(w[10], 10, format_num(data['sgst']), 1, 0, 'R')
        
    pdf.cell(w[11], 10, "-", 1, 0, 'C')
    pdf.cell(w[12], 10, "-", 1, 0, 'R')
    pdf.cell(w[13], 10, format_num(data['total_amount']), 1, 1, 'R')
    
    pdf.cell(w[0], 8, "Airport Charges", 1, 0, 'L')
    pdf.cell(w[1], 8, "", 1, 0, 'C')
    pdf.cell(w[2], 8, "", 1, 0, 'C')
    pdf.cell(w[3], 8, "-", 1, 0, 'R')
    pdf.cell(w[4], 8, "-", 1, 0, 'R')
    for i in range(5, 13): pdf.cell(w[i], 8, "", 1, 0)
    pdf.cell(w[13], 8, "-", 1, 1, 'R')
    
    pdf.set_font('helvetica', 'B', 9)
    pdf.cell(w[0]+w[1], 8, "Grand Total", 1, 0, 'L')
    pdf.cell(w[2], 8, format_num(data['taxable_value']), 1, 0, 'R')
    pdf.cell(w[3], 8, "-", 1, 0, 'R')
    pdf.cell(w[4], 8, format_num(data['taxable_value']), 1, 0, 'R')
    
    pdf.cell(w[5], 8, "", 1, 0)
    pdf.cell(w[6], 8, format_num(igst) if igst>0 else "-", 1, 0, 'R')
    pdf.cell(w[7], 8, "", 1, 0)
    pdf.cell(w[8], 8, format_num(data['cgst']) if float(data['cgst'] or 0)>0 else "-", 1, 0, 'R')
    pdf.cell(w[9], 8, "", 1, 0)
    pdf.cell(w[10], 8, format_num(data['sgst']) if float(data['sgst'] or 0)>0 else "-", 1, 0, 'R')
    pdf.cell(w[11], 8, "", 1, 0)
    pdf.cell(w[12], 8, "-", 1, 0, 'R')
    pdf.cell(w[13], 8, format_num(data['total_amount']), 1, 1, 'R')
    
    pdf.cell(sum(w[:5]), 8, "Total Invoice Amount (in figures)", 1, 0, 'L')
    pdf.set_font('helvetica', 'B', 9)
    pdf.cell(sum(w[5:]), 8, f" {str(data['amount_in_words'])}", 1, 1, 'L')
    
    pdf.ln(5)
    pdf.set_font('helvetica', '', 8)
    notes = [
        "1. Air Travel And Related Charges :- Includes all charges related to air transportation of passengers.",
        "2. Airport Charges :- Includes ADF, UDF, PSF and other airport charges collected on behalf of Airport operator, as applicable",
        "3. Meal :- Includes all prepaid meals purchased before travel",
        "4. Amounts have been rounded off.",
        "5. I/We hereby declare that though our aggregate turnover in any preceding financial year from 2017-18 onwards is more than the aggregate turnover notified under sub-rule (4) of rule 48, we are not required to prepare an invoice in terms of the provisions of the said sub-rule"
    ]
    cur_y = pdf.get_y()
    for note in notes:
        pdf.set_x(10)
        pdf.multi_cell(180, 4, note)
    
    pdf.set_y(cur_y)
    pdf.cell(0, 5, "For Just Udo Aviation Pvt. Ltd.", 0, 1, 'R')
    y_sig = pdf.get_y()
    if seal_pos and os.path.exists(SEAL_PATH):
        pdf.image(SEAL_PATH, seal_pos.get('x', 248), seal_pos.get('y', y_sig - 5), seal_pos.get('w', 30))
    if sign_pos and os.path.exists(SIGN_PATH):
        pdf.image(SIGN_PATH, sign_pos.get('x', 250), sign_pos.get('y', y_sig + 1), sign_pos.get('w', 25))
    pdf.set_y(y_sig + 15)
    pdf.cell(0, 5, "Authorised Signatory", 0, 1, 'R')
    pdf.output(output_path)

def process_all_invoices():
    data_rows = excel_to_dict_list(EXCEL_FILE, 'Data')
    address_rows = excel_to_dict_list(EXCEL_FILE, 'Address Master')
    fly91_rows = excel_to_dict_list(EXCEL_FILE, 'FLY91 Address Master')
    
    customer_address_lookup = {str(r.get('GSTIN')): str(r.get('Address')) for r in address_rows if not is_empty(r.get('GSTIN'))}
    supplier_address_lookup = {str(r.get('GSTIN')): str(r.get('Address')) for r in fly91_rows if not is_empty(r.get('GSTIN'))}

    if not os.path.exists(OUTPUT_BASE_DIR): os.makedirs(OUTPUT_BASE_DIR)

    for index, row in enumerate(data_rows[:50]):
        invoice_no = str(row.get('Invoicenumber'))
        if is_empty(invoice_no): continue
        
        supplier_gstin = str(row.get('FLY91 GSTIN', ''))
        supplier_address = supplier_address_lookup.get(supplier_gstin, "Address not found")
        customer_gstin = str(row.get('GSTIN', '-'))
        customer_address = customer_address_lookup.get(customer_gstin, "-")
        
        raw_date = row.get('Invoice Date')
        if isinstance(raw_date, datetime.datetime):
            inv_date = raw_date.strftime('%d-%m-%Y')
        else:
            inv_date = str(raw_date)[:10] if not is_empty(raw_date) else ""
            
        total_val = float(row.get('Invoice Value', 0)) if not is_empty(row.get('Invoice Value')) else 0.0
        rounded_total_val = round(total_val)
        
        flight_no = str(int(round(float(row.get('Flight Number'))))) if not is_empty(row.get('Flight Number')) else ""
        sac_code = str(int(round(float(row.get('HSN'))))) if not is_empty(row.get('HSN')) else "996425"
        
        data = {
            'supplier_address': supplier_address,
            'supplier_gstin': supplier_gstin,
            'invoice_no': invoice_no, 'invoice_date': inv_date,
            'passenger_name': str(row.get('Passenger Name', '')).upper(),
            'pnr_no': str(row.get('PNRNumber', '')), 'flight_no': flight_no,
            'origin': str(row.get('Origin', '')), 'destination': str(row.get('Destination', '')),
            'place_of_supply': str(row.get('Place of supply - State', '')).upper(),
            'customer_name': str(row.get('Customer Name ', '')).upper(),
            'customer_address': customer_address, 'customer_gstin': customer_gstin,
            'description': str(row.get('DESCRIPTION ON INVOICE', 'Airport travel and related charges')),
            'sac_code': sac_code, 'taxable_value': float(row.get('Taxable Value', 0)) if not is_empty(row.get('Taxable Value')) else 0.0,
            'igst': float(row.get('IGST', 0)) if not is_empty(row.get('IGST')) else 0.0,
            'cgst': float(row.get('CGST', 0)) if not is_empty(row.get('CGST')) else 0.0,
            'sgst': float(row.get('SGST', 0)) if not is_empty(row.get('SGST')) else 0.0,
            'total_amount': total_val, 'amount_in_words': number_to_words_indian(rounded_total_val)
        }
        
        folder_name = clean_filename(str(row.get('Folder bifurcation', 'Unknown')))
        target_dir = os.path.join(OUTPUT_BASE_DIR, folder_name)
        if not os.path.exists(target_dir): os.makedirs(target_dir)
        output_pdf = os.path.join(target_dir, f"{clean_filename(invoice_no)}.pdf")
        generate_kind_pdf(data, output_pdf)

def get_excel_data_rows(excel_path):
    rows = excel_to_dict_list(excel_path, 'Data')
    # Filter valid invoice rows
    valid_rows = [r for r in rows if not is_empty(r.get('Invoicenumber'))]
    return valid_rows

def get_lookups(excel_path):
    address_rows = excel_to_dict_list(excel_path, 'Address Master')
    fly91_rows = excel_to_dict_list(excel_path, 'FLY91 Address Master')
    return {
        'customer': {str(r.get('GSTIN')): str(r.get('Address')) for r in address_rows if not is_empty(r.get('GSTIN'))},
        'supplier': {str(r.get('GSTIN')): str(r.get('Address')) for r in fly91_rows if not is_empty(r.get('GSTIN'))}
    }

def get_invoicing_data(row, lookups):
    invoice_no = str(row.get('Invoicenumber'))
    supplier_gstin = str(row.get('FLY91 GSTIN', ''))
    supplier_address = lookups['supplier'].get(supplier_gstin, "Address not found")
    customer_gstin = str(row.get('GSTIN', '-'))
    customer_address = lookups['customer'].get(customer_gstin, "-")
    
    raw_date = row.get('Invoice Date')
    if isinstance(raw_date, datetime.datetime):
        inv_date = raw_date.strftime('%d-%m-%Y')
    else:
        inv_date = str(raw_date)[:10] if not is_empty(raw_date) else ""
    
    total_val = float(row.get('Invoice Value', 0)) if not is_empty(row.get('Invoice Value')) else 0.0
    rounded_total_val = round(total_val)
    
    flight_no = str(int(round(float(row.get('Flight Number'))))) if not is_empty(row.get('Flight Number')) else ""
    sac_code = str(int(round(float(row.get('HSN'))))) if not is_empty(row.get('HSN')) else "996425"
    
    return {
        'supplier_address': supplier_address, 'supplier_gstin': supplier_gstin,
        'invoice_no': invoice_no, 'invoice_date': inv_date,
        'passenger_name': str(row.get('Passenger Name', '')).upper(),
        'pnr_no': str(row.get('PNRNumber', '')), 'flight_no': flight_no,
        'origin': str(row.get('Origin', '')), 'destination': str(row.get('Destination', '')),
        'place_of_supply': str(row.get('Place of supply - State', '')).upper(),
        'customer_name': str(row.get('Customer Name ', '')).upper(),
        'customer_address': customer_address, 'customer_gstin': customer_gstin,
        'description': str(row.get('DESCRIPTION ON INVOICE', 'Airport travel and related charges')),
        'sac_code': sac_code, 'taxable_value': float(row.get('Taxable Value', 0)) if not is_empty(row.get('Taxable Value')) else 0.0,
        'igst': float(row.get('IGST', 0)) if not is_empty(row.get('IGST')) else 0.0,
        'cgst': float(row.get('CGST', 0)) if not is_empty(row.get('CGST')) else 0.0,
        'sgst': float(row.get('SGST', 0)) if not is_empty(row.get('SGST')) else 0.0,
        'total_amount': total_val, 'amount_in_words': number_to_words_indian(rounded_total_val),
        'folder_bifurcation': clean_filename(str(row.get('Folder bifurcation', 'Unknown')))
    }

if __name__ == "__main__":
    process_all_invoices()
    print("Horizontal Layout complete.")
